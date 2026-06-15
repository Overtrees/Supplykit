from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException
from datetime import datetime
import json, csv, io, re, os, uuid
from openpyxl import load_workbook
from app.core.database import get_db, submit_task, get_task, backup_db
from app.api.routes.ws import broadcast
from app.api.routes.insights import auto_adjust_inventory

router = APIRouter(prefix="/api/cleansing", tags=["cleansing"])

# ─── 系统目标字段定义 ────────────────────────────────────────────────────────

SYSTEM_FIELDS = {
    'order': [
        {'key':'order_no',       'label':'订单号',     'type':'string'},
        {'key':'store',          'label':'店铺',       'type':'string'},
        {'key':'warehouse',      'label':'仓库',       'type':'string'},
        {'key':'sku',            'label':'商品编号',   'type':'string'},
        {'key':'product_name',   'label':'商品名称',   'type':'string'},
        {'key':'quantity',       'label':'数量',       'type':'number'},
        {'key':'unit_price',     'label':'单价',       'type':'number'},
        {'key':'total_amount',   'label':'总金额',     'type':'number'},
        {'key':'order_status',   'label':'状态',       'type':'string'},
        {'key':'ordered_at',     'label':'日期',       'type':'date'},
        {'key':'supplier',       'label':'供应商',     'type':'string'},
        {'key':'remark',         'label':'备注',       'type':'string'},
    ],
    'inventory': [
        {'key':'store',          'label':'店铺',       'type':'string'},
        {'key':'warehouse',      'label':'仓库',       'type':'string'},
        {'key':'sku',            'label':'商品编号',   'type':'string'},
        {'key':'product_name',   'label':'商品名称',   'type':'string'},
        {'key':'available_qty',  'label':'可用库存',   'type':'number'},
        {'key':'locked_qty',     'label':'锁定库存',   'type':'number'},
        {'key':'in_transit_qty', 'label':'在途数量',   'type':'number'},
        {'key':'safety_qty',     'label':'安全库存',   'type':'number'},
    ],
}

# ─── 自定义字段存储 ────────────────────────────────────────────────────────────

CUSTOM_FIELDS_PATH = '/home/Overtrees/Supplykit/backend/custom_fields.json'

def load_custom_fields():
    if os.path.exists(CUSTOM_FIELDS_PATH):
        try:
            with open(CUSTOM_FIELDS_PATH) as f:
                return json.load(f)
        except: pass
    return {'order': [], 'inventory': []}

def save_custom_fields(data):
    os.makedirs(os.path.dirname(CUSTOM_FIELDS_PATH), exist_ok=True)
    with open(CUSTOM_FIELDS_PATH, 'w') as f:
        json.dump(data, f, ensure_ascii=False)

# ─── 文件解析 ────────────────────────────────────────────────────────────────

def parse_file(content, filename):
    if filename.lower().endswith('.csv'):
        text = content.decode('utf-8-sig', errors='ignore')
        return list(csv.DictReader(io.StringIO(text)))
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[wb.sheetnames[0]]
    raw = list(ws.iter_rows(values_only=True))
    if not raw:
        return []
    headers = [str(c).strip() if c is not None else '' for c in raw[0]]
    return [{headers[i]: raw[r][i] for i in range(len(headers))} for r in range(1, len(raw))]

# ─── 检测接口 ────────────────────────────────────────────────────────────────

@router.post('/detect')
async def detect_columns(file: UploadFile = File(...)):
    content = await file.read()
    rows = parse_file(content, file.filename)
    if not rows:
        return {'ok': False, 'error': '文件为空'}
    cols = []
    for key in rows[0].keys():
        samples = []
        for r in rows[:5]:
            v = r.get(key)
            if v is not None and str(v).strip():
                samples.append(str(v)[:60])
        cols.append({'name': key, 'samples': samples[:3], 'count': len(rows)})
    return {'ok': True, 'columns': cols, 'total': len(rows), 'file': file.filename}

# ─── 预览接口 ────────────────────────────────────────────────────────────────

@router.post('/preview')
async def preview_cleansing(file: UploadFile = File(...), mapping: str = Form('')):
    content = await file.read()
    rows = parse_file(content, file.filename)
    if not rows:
        return {'ok': False, 'error': '文件为空'}
    try:
        mapping_config = json.loads(mapping) if mapping else {}
    except json.JSONDecodeError:
        return {'ok': False, 'error': '映射配置格式错误'}

    preview_rows = []
    for row in rows[:20]:
        result = {'_source': {}}
        for src_col, cfg in mapping_config.items():
            target = cfg.get('target', '')
            if not target:
                continue
            raw_val = row.get(src_col, '')
            cleaned = cleanse_value(raw_val, cfg)
            result['_source'][src_col] = str(raw_val)[:80] if raw_val is not None else ''
            result[target] = cleaned
        preview_rows.append(result)

    return {'ok': True, 'preview': preview_rows, 'total': len(rows), 'mapped': len(mapping_config)}

# ─── 执行清洗 ────────────────────────────────────────────────────────────────

def _run_cleansing(content: bytes, filename: str, mapping_json: str, target: str, template_name: str = ''):
    """清洗核心逻辑，含格式校验 → 业务校验 → 补全推断"""
    db = get_db()
    rows = parse_file(content, filename)
    if not rows:
        return {'ok': False, 'error': '文件为空', 'success': 0, 'failed': 0, 'file': filename}
    try:
        mapping_config = json.loads(mapping_json) if mapping_json else {}
    except json.JSONDecodeError:
        return {'ok': False, 'error': '映射配置格式错误', 'success': 0, 'failed': 0, 'file': filename}

    # 加载用于校验和推断的参考数据
    products_map = {p["sku"]: p for p in db.table("products").select("*").execute().data}
    inventory_map = {i["sku"]: i for i in db.table("inventory").select("*").execute().data}
    task_id = f"clean_{datetime.utcnow().strftime('%H%M%S')}"
    errors = []
    success = 0
    failed = 0
    orders_to_insert = []
    order_no_seen = set()
    dedup = {}

    for idx, row in enumerate(rows):
        row_errors = []
        data = {}

        # ─── 格式校验 + 字段映射 ──────────────────────────────────────
        for src_col, cfg in mapping_config.items():
            target_field = cfg.get('target', '')
            if not target_field: continue
            raw_val = row.get(src_col, '')
            try:
                cleaned = cleanse_value(raw_val, cfg)
                data[target_field] = cleaned
            except Exception as e:
                row_errors.append({'error_type': 'format_error', 'field_name': src_col,
                                   'raw_value': str(raw_val)[:100], 'error_message': str(e)[:100]})

        # ─── 补全推断 ──────────────────────────────────────────────
        sku = str(data.get('sku', ''))
        if sku and not data.get('product_name'):
            p = products_map.get(sku)
            if p:
                data['product_name'] = p.get('product_name', '')

        # ─── 业务校验 ──────────────────────────────────────────────
        if not data.get('ordered_at'):
            data['ordered_at'] = datetime.utcnow().strftime('%Y-%m-%d')

        # 订单号处理
        order_no = data.get('order_no', '')
        if not order_no:
            order_no = f"AUTO-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}-{success}"
            data['order_no'] = order_no
        if order_no in dedup:
            dedup[order_no] += 1
            order_no = f"{order_no}-{dedup[order_no]}"
        else:
            dedup[order_no] = 0
        data['order_no'] = order_no

        # 记录行错误（不影响继续处理，只是标记）
        for e in row_errors:
            errors.append(e)
            try:
                db.table("cleansing_errors").insert({
                    "task_id": task_id, "row_index": idx, "source_file": filename,
                    "error_type": e['error_type'], "field_name": e['field_name'],
                    "raw_value": e['raw_value'], "error_message": e['error_message'],
                    "raw_data": json.dumps(row, ensure_ascii=False, default=str),
                }).execute()
            except: pass

        # ─── 写入目标 ──────────────────────────────────────────────
        if order_no not in order_no_seen:
            order_no_seen.add(order_no)
            orders_to_insert.append({
                "order_no": order_no, "store": str(data.get('store', '未知'))[:100],
                "sku": sku[:100],
                "product_name": str(data.get('product_name', ''))[:200],
                "quantity": int(float(data.get('quantity', 0))),
                "unit_price": float(data.get('unit_price', 0)),
                "total_amount": float(data.get('total_amount', 0)),
                "order_status": str(data.get('order_status', '已完成'))[:50],
                "ordered_at": str(data.get('ordered_at', ''))[:50],
            })
            success += 1
        else:
            failed += 1
            errors.append({'error_type': 'duplicate_order', 'field_name': 'order_no',
                           'raw_value': order_no, 'error_message': '重复订单号'})

    if orders_to_insert:
        try:
            db.table("orders").insert(orders_to_insert).execute()
            # 触犯事件
            try:
                from app.core.events import bus
                bus.emit('data.cleaned', {
                    'target': target, 'event_type': f'{target}.cleansed',
                    'title': f'清洗导入 {success} 条',
                    'success': success, 'failed': failed,
                    'ws_message': {
                        'type': f'{target}.cleansed',
                        'payload': {'success': success, 'failed': failed}
                    }
                })
            except: pass
            from app.core.database import submit_task
            from app.api.routes.insights import sync_inventory_from_orders
            submit_task(f"inv_sync_{datetime.utcnow().strftime('%H%M%S')}", sync_inventory_from_orders, 200)
        except Exception as e:
            return {'ok': False, 'error': f'清洗写入失败: {str(e)[:200]}', 'success': 0, 'failed': 0}

    msg_parts = []
    if success > 0: msg_parts.append(f"成功导入 {success} 条")
    if failed > 0: msg_parts.append(f"{failed} 条跳过")
    if errors: msg_parts.append(f"{len(errors)} 条异常（可查看错误详情）")

    if success == 0 and failed > 0:
        return {'ok': False, 'success': 0, 'failed': failed, 'file': filename, 'target': target,
                'error': '所有记录均重复或写入失败', 'error_count': len(errors)}
    return {'ok': True, 'success': success, 'failed': failed, 'file': filename,
            'target': target, 'message': '，'.join(msg_parts) if msg_parts else '无数据变更',
            'error_count': len(errors)}

@router.get('/errors')
def get_cleansing_errors(file: str = '', db = get_db()):
    """查询清洗错误记录"""
    if file:
        errs = db.table("cleansing_errors").select("*").eq("source_file", file).order("id", desc=True).limit(500).execute().data
    else:
        errs = db.table("cleansing_errors").select("*").order("id", desc=True).limit(200).execute().data
    for e in errs:
        try: e['raw_data'] = json.loads(e.get('raw_data','{}'))
        except: pass
    return {'ok': True, 'errors': errs, 'total': len(errs)}

@router.post('/execute')
async def execute_cleansing(file: UploadFile = File(...), mapping: str = Form(''),
                             target: str = Form('order'), template_name: str = Form('')):
    content = await file.read()
    return _run_cleansing(content, file.filename, mapping, target, template_name)
@router.get('/templates')
def list_templates(db = get_db()):
    templates = db.table("cleansing_templates").select("*").order("updated_at", desc=True).execute().data
    return [{
        'id': t['id'], 'name': t['name'], 'doc_type': t['doc_type'],
        'mapping': json.loads(t.get('mapping') or '{}'),
        'updated_at': t.get('updated_at'),
    } for t in templates]

@router.post('/templates')
def save_template(data: dict, db = get_db()):
    name = data.get('name', '').strip()
    if not name:
        raise HTTPException(status_code=400, detail='模板名称不能为空')
    existing = db.table("cleansing_templates").select("id").eq("name", name).execute().data
    payload = {
        'name': name,
        'doc_type': data.get('doc_type', 'order'),
        'mapping': json.dumps(data.get('mapping', {}), ensure_ascii=False),
        'updated_at': datetime.utcnow().isoformat()[:19],
    }
    if existing:
        db.table("cleansing_templates").update(payload).eq("id", existing[0]['id']).execute()
        return {'ok': True, 'message': f'模板「{name}」已更新', 'id': existing[0]['id']}
    else:
        db.table("cleansing_templates").insert(payload).execute()
        return {'ok': True, 'message': f'模板「{name}」已保存'}

@router.delete('/templates/{template_id}')
def delete_template(template_id: int, db = get_db()):
    data = db.table("cleansing_templates").select("id").eq("id", template_id).execute().data
    if not data:
        raise HTTPException(status_code=404, detail='模板不存在')
    db.table("cleansing_templates").delete().eq("id", template_id).execute()
    return {'ok': True}

# ─── 字段管理 ────────────────────────────────────────────────────────────────

@router.get('/fields/{target}')
def get_fields(target: str):
    system = SYSTEM_FIELDS.get(target)
    if not system:
        raise HTTPException(status_code=404, detail=f'目标 {target} 不存在')
    custom = load_custom_fields().get(target, [])
    return {'system': system, 'custom': custom, 'all': system + custom}

@router.get('/custom-fields/{target}')
def list_custom_fields(target: str):
    data = load_custom_fields()
    return data.get(target, [])

@router.post('/custom-fields/{target}')
def add_custom_field(target: str, data: dict):
    if target not in ('order', 'inventory'):
        raise HTTPException(status_code=400, detail='目标必须是 order 或 inventory')
    key = str(data.get('key', '')).strip()
    label = str(data.get('label', key)).strip()
    ftype = str(data.get('type', 'string')).strip()
    if not key:
        raise HTTPException(status_code=400, detail='字段名不能为空')
    cf = load_custom_fields()
    existing = [f for f in cf.get(target, []) if f['key'] == key]
    if not existing:
        cf[target].append({'key': key, 'label': label, 'type': ftype})
        save_custom_fields(cf)
    return {'ok': True, 'fields': cf[target]}

@router.delete('/custom-fields/{target}/{field_key}')
def remove_custom_field(target: str, field_key: str):
    cf = load_custom_fields()
    cf[target] = [f for f in cf.get(target, []) if f['key'] != field_key]
    save_custom_fields(cf)
    return {'ok': True}

@router.post('/execute-async')
async def execute_cleansing_async(file: UploadFile = File(...), mapping: str = Form(''),
                                   target: str = Form('order'), template_name: str = Form('')):
    content = await file.read()
    task_id = str(uuid.uuid4())[:8]
    submit_task(task_id, _run_cleansing, content, file.filename, mapping, target, template_name)
    return {'ok': True, 'task_id': task_id, 'message': '任务已提交'}

# ─── 异步任务进度 ───────────────────────────────────────────────────────────

@router.get('/task/{task_id}')
def get_task_status(task_id: str):
    task = get_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail='任务不存在')
    return {'ok': True, 'task_id': task_id, **task}

# ─── 数据库备份 ──────────────────────────────────────────────────────────────

@router.post('/backup')
def trigger_backup():
    path = backup_db()
    if path:
        return {'ok': True, 'path': path, 'message': f'备份完成: {os.path.basename(path)}'}
    return {'ok': False, 'error': '备份失败'}

# ─── 清洗工具函数 ────────────────────────────────────────────────────────────

def cleanse_value(raw_val, cfg):
    if raw_val is None or str(raw_val).strip() == '':
        return cfg.get('default', '')
    v = str(raw_val).strip()
    field_type = cfg.get('type', 'string')
    fmt_str = cfg.get('format', '')
    try:
        if field_type == 'number':
            cleaned = re.sub(r'[^\d.\-]', '', v)
            return float(cleaned) if '.' in cleaned else int(float(cleaned))
        elif field_type == 'date':
            if fmt_str == 'YMD':
                return v[:10]
            return v
        else:
            return v
    except:
        return v
