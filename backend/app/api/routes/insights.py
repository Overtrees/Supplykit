from fastapi import APIRouter, Depends
from app.core.database import get_db
from app.core.response import ok, fail
from app.core.sales_utils import calc_sales, rolling_predict
from app.api.routes.replenishment import get_replenishment_suggestions
from datetime import datetime
import json, os

router = APIRouter(prefix="/api/insights", tags=["insights"])

# with-sales 结果缓存（30s TTL + 版本号）
_with_sales_cache = {}


@router.get('/ping')
def ping():
    return ok({"time": datetime.utcnow().isoformat()})

def detect_slow_moving_products(db=None, create_alerts=False):
    from datetime import datetime, timedelta
    if db is None:
        from app.core.database import get_db
        db = get_db()
    cutoff = (datetime.utcnow() - timedelta(days=90)).isoformat()
    # 快照聚合：按 SKU 取最大日期，替代 orders 全表 GROUP BY
    from app.core.database import get_conn
    last_order = {}
    try:
        rows = get_conn().execute("SELECT sku, MAX(date) FROM daily_sales_snapshot WHERE date >= ? GROUP BY sku", (cutoff,)).fetchall()
        last_order = {r[0]: (r[1] or '')[:10] for r in rows}
    except Exception as e:
        import logging; logging.warning(f"[slow-moving] snapshot agg: {e}")
    # 当天 orders 补充（快照不含今天）
    try:
        today = datetime.utcnow().strftime('%Y-%m-%d')
        rows = get_conn().execute("SELECT sku, MAX(ordered_at) FROM orders WHERE ordered_at >= ? GROUP BY sku", (today,)).fetchall()
        for r in rows:
            if r[0] and (r[1] or '')[:10] > last_order.get(r[0], ''):
                last_order[r[0]] = (r[1] or '')[:10]
    except Exception as e:
        import logging; logging.warning(f"[slow-moving] today orders: {e}")
    # 只加载需要的字段，避免全量 select("*") 导致 10 万 SKU 时 OOM
    products_map = {}
    try:
        from app.core.database import get_conn
        _conn = get_conn()
        for r in _conn.execute("SELECT sku, product_name, barcode, channel FROM products WHERE deleted_at IS NULL OR deleted_at=''").fetchall():
            products_map[r[0]] = {"sku": r[0], "product_name": r[1], "barcode": r[2] or '', "channel": r[3] or 'jd'}
    except Exception as e:
        import logging; logging.warning(f"[slow-moving] products: {e}")
    sku_barcode_map = {s: (p.get('barcode', '') or '') for s, p in products_map.items()}
    inventory_map = {}
    try:
        for r in _conn.execute("SELECT sku, available_qty, product_name, channel FROM inventory").fetchall():
            inventory_map[r[0]] = {"sku": r[0], "available_qty": r[1], "product_name": r[2] or r[0], "channel": r[3] or 'jd'}
    except Exception as e:
        import logging; logging.warning(f"[slow-moving] inventory: {e}")
    # SKU → channel（优先 products 主表，回退 inventory）
    from app.core.sales_utils import sku_to_channel
    sku_channel_map = {s: (p.get('channel') or sku_to_channel(s, db) or 'jd') for s, p in products_map.items()}
    for s, i in inventory_map.items():
        if s not in sku_channel_map or not sku_channel_map[s]:
            sku_channel_map[s] = i.get('channel') or 'jd'
    now = datetime.utcnow()
    result = []
    # 只遍历有库存的 SKU（无库存不需要滞销检测），避免 10 万+ SKU 全量遍历
    all_skus = set(inventory_map.keys())
    for sku in all_skus:
        p = products_map.get(sku)
        inv = inventory_map.get(sku)
        bc = sku_barcode_map.get(sku, '')
        key = f"{sku}|{bc}" if bc else sku
        last_date = last_order.get(key, "")
        days = 999
        if last_date:
            try: days = (now - datetime.strptime(last_date[:10], "%Y-%m-%d")).days
            except Exception as e: import logging; logging.warning(f"[slow-moving] parse date {last_date}: {e}")
        stock = int(inv.get("available_qty") or 0) if inv else 0
        if days > 30 and stock > 0:
            level = "滞销" if days > 60 else ("冷淡" if days > 30 else "正常")
            result.append({"sku": sku, "barcode": bc, "product_name": p["product_name"] if p else inv.get("product_name",sku) if inv else sku, "last_order_date": last_date[:10], "days_since_last": days, "stock": stock, "level": level, "channel": sku_channel_map.get(sku, 'jd')})
            if create_alerts:
                ex = db.table("alerts").select("id").eq("alert_type","slow_moving").eq("related_sku",sku).eq("status","active").execute().data
                if not ex:
                    db.table("alerts").insert({"alert_type":"slow_moving", "title":f"滞销: {result[-1]['product_name']}", "description":f"{days} 天无销售，库存 {stock} 件", "severity":"warning", "source":"event_bus", "related_sku":sku, "status":"active", "channel": sku_channel_map.get(sku, 'jd')}).execute()
    result.sort(key=lambda x: -x["days_since_last"])
    return ok(result)

@router.get('/slow-moving')
def get_slow_moving_products(db = get_db()):
    return detect_slow_moving_products(db, create_alerts=False)


@router.get('/disposal-suggestions')
def get_disposal_suggestions(channel: str = 'jd', db = get_db()):
    """滞销品自动处置建议 v6 — 组合判定（SKU×仓库粒度）

    维度（全部可配，参数页「滞销参数」tab）：
    - 时间: 连续零销售天数（最后销售日-今天）
    - 动销率: 近N天销售件数/当前库存（食品45天<8%, 家清60天<3%）
    - 周转/存销比: 库存/28天日均销
    - 深度: SKU库存/近90天总销量占比 + 连续无订单天数
    - 临期: best_before 距今天数（食品<3月/个护<6月=高风险）
    - B仓仓储费: 超免费期天数×占用方数×费率（京东BBCC）
    - ABC: 近90天销售额分档（A核心不轻处置 / C类重点监控）
    等级: 临期(black) > 深度积压(deep) > 确认滞销(confirm) > 潜在(potential) > 预警(warn)
    """
    from datetime import datetime, timedelta
    from app.core.database import get_conn
    from app.core.sales_utils import load_daily_sales, calc_sales_from_daily
    conn = get_conn()
    today = datetime.utcnow()
    today_s = today.strftime('%Y-%m-%d')

    # ── 配置读取（与滞销参数页同 key）──
    def cfg(key, default):
        try:
            r = conn.execute("SELECT value FROM replenishment_config WHERE key=? AND channel=?", (key, channel)).fetchone()
            return r[0] if r and r[0] else default
        except Exception:
            return default
    FC = {  # food config / nonfood config 简化: 先按品类取通用值
        'alert_days': int(cfg('slow_alert_days', 14)),
        'pot_food': int(cfg('slow_potential_food', 30)), 'pot_nonfood': int(cfg('slow_potential_nonfood', 60)),
        'conf_food': int(cfg('slow_confirm_food', 60)), 'conf_nonfood': int(cfg('slow_confirm_nonfood', 90)),
        'ratio_confirm': float(cfg('slow_ratio_confirm', 2.5)),
        'turn_food': float(cfg('slow_turnover_food', 45)), 'turn_nonfood': float(cfg('slow_turnover_nonfood', 60)),
        'rate_food': float(cfg('slow_turnrate_food', 8)), 'rate_nonfood': float(cfg('slow_turnrate_nonfood', 3)),
        'window': int(cfg('slow_turn_window', 45)),
        'deep_ratio': float(cfg('slow_deep_ratio', 85)), 'deep_noorder': int(cfg('slow_deep_noorder', 14)),
        'shelf_food': int(cfg('slow_shelf_food', 3)), 'shelf_nonfood': int(cfg('slow_shelf_nonfood', 6)),
        'abc_a': float(cfg('abc_a_ratio', 20)) / 100, 'abc_b': float(cfg('abc_b_ratio', 50)) / 100,
        'b_free': int(cfg('b_free_days', 15)), 'fee_rate': float(cfg('b_storage_fee_rate', 1.0)),
    }

    # ── 品类映射：调味+零食=食品类 / 日化=家清类 ──
    FOOD_KEYWORDS = ('酱油','酱','醋','油','料酒','蚝油','辣椒','调味','花椒','藤椒','芥末','番茄','沙拉',
                     '芝麻','花生','豆瓣','豆豉','腐乳','糟卤','鱼露','咖喱','五香','孜然','胡椒','十三香',
                     '卤料','炖肉','鸡精','味精','白糖','冰糖','红糖','麦芽糖','蜂蜜','黄酒','米酒','薯片',
                     '虾条','爆米花','坚果','瓜子','饼干','威化','巧克力','糖果')

    # 商品信息（价格/体积/品类/保质期）
    products = {}
    try:
        for r in conn.execute("SELECT sku, product_name, price, volume, category, best_before, channel FROM products WHERE (deleted_at IS NULL OR deleted_at='') AND channel=?", (channel,)).fetchall():
            cat = str(r[4] or '')
            is_food = cat.lower() in ('food','食品') or any(k in cat for k in FOOD_KEYWORDS)
            products[str(r[0])] = {"name": str(r[1] or ''), "price": float(r[2] or 0), "volume": float(r[3] or 0),
                              "food": is_food, "best_before": str(r[5] or '')[:10]}
    except Exception as e:
        import logging; logging.warning(f"[disposal] products: {e}")

    # 日销（28 天）
    try:
        sales_28 = calc_sales_from_daily(load_daily_sales(28, db, channel=channel), 28)
    except Exception as e:
        import logging; logging.warning(f"[disposal] sales: {e}")
        sales_28 = {}
    # 近 N 天销量（动销率分子）与近 90 天（深度分母 + ABC 销售额）
    sale_90 = {}; sale_win = {}; sale_days = {}  # sku -> {total, days_with_sale}
    try:
        cutoff90 = (today - timedelta(days=90)).strftime('%Y-%m-%d')
        cutoff_w = (today - timedelta(days=FC['window'])).strftime('%Y-%m-%d')
        rows = conn.execute("SELECT sku, ordered_at, quantity, total_amount FROM orders WHERE channel=? AND ordered_at>=?", (channel, cutoff90)).fetchall()
        for r in rows:
            sk = str(r[0]); dt = str(r[1])[:10]; qty = int(r[2] or 0)
            amt = float(r[3] or 0)
            s90 = sale_90.setdefault(sk, {'qty': 0, 'amt': 0, 'days': set()})
            s90['qty'] += qty; s90['amt'] += amt; s90['days'].add(dt)
            if dt >= cutoff_w:
                sw = sale_win.setdefault(sk, 0); sale_win[sk] = sw + qty
    except Exception as e:
        import logging; logging.warning(f"[disposal] order agg: {e}")
    # 最后销售日（快照 MAX + sale_90 已遍历的当天/近90天订单，避免重复全表查询）
    last_order = {}
    try:
        for r in conn.execute("SELECT sku, MAX(date) FROM daily_sales_snapshot WHERE channel=? GROUP BY sku", (channel,)).fetchall():
            last_order[str(r[0])] = str(r[1] or '')[:10]
    except Exception as e:
        import logging; logging.warning(f"[disposal] snapshot date: {e}")
    for sk, v in sale_90.items():
        if v['days']:
            _mx = max(v['days'])
            if _mx > last_order.get(sk, ''):
                last_order[sk] = _mx
    # ABC 分档（近90天销售额）
    abc_map = {}
    try:
        ranked = sorted(sale_90.items(), key=lambda kv: -kv[1]['amt'])
        total_amt = sum(v['amt'] for _, v in ranked) or 1
        cum = 0.0
        for sk, v in ranked:
            cum += v['amt'] / total_amt
            abc_map[sk] = 'a' if cum <= FC['abc_a'] else ('b' if cum <= FC['abc_b'] else 'c')
    except Exception as e:
        import logging; logging.warning(f"[disposal] abc: {e}")
    # B 仓入库批次
    b_arrival = {}
    try:
        for r in conn.execute("SELECT sku, arrival_date FROM purchase_orders WHERE channel=? AND arrival_date != ''", (channel,)).fetchall():
            if r[1]:
                try:
                    d = datetime.strptime(str(r[1])[:10], "%Y-%m-%d")
                    b_arrival.setdefault(str(r[0]), d)
                except Exception: pass
    except Exception as e:
        import logging; logging.warning(f"[disposal] po: {e}")
    # 已处置记录（30 天去重）
    disposed = {}
    try:
        cutoff30 = (today - timedelta(days=30)).strftime('%Y-%m-%d %H:%M:%S')

        for r in conn.execute("SELECT sku, warehouse, action FROM disposal_records WHERE channel=? AND created_at >= ?", (channel, cutoff30)).fetchall():
            disposed[(str(r[0]), str(r[1]))] = str(r[2] or '')
    except Exception as e:
        import logging; logging.warning(f"[disposal] disposed: {e}")
    # 近14天补货建议（伪滞销线索）
    pseudo = set()
    try:
        c14 = (today - timedelta(days=14)).strftime('%Y-%m-%d %H:%M:%S')
        for r in conn.execute("SELECT DISTINCT related_sku FROM alerts WHERE channel=? AND alert_type IN ('replenish','rewrite') AND created_at>=?", (channel, c14)).fetchall():
            if r[0]: pseudo.add(str(r[0]))
    except Exception as e:
        import logging; logging.warning(f"[disposal] pseudo: {e}")

    suggestions = []
    LEVEL_ORDER = {'black': 0, 'deep': 1, 'confirm': 2, 'potential': 3, 'warn': 4}
    SUG = {'black': '临期商品紧急处理(退供/促销)', 'deep': '深度积压: 退货/清仓', 'confirm': '确认滞销: 退货供应商/清仓甩卖',
           'potential': '潜在滞销: 降价促销/调整策略', 'warn': '预警: 补货降量/持续跟踪'}
    for r in conn.execute("SELECT sku, warehouse, warehouse_type, available_qty, in_transit_qty FROM inventory WHERE channel=?", (channel,)).fetchall():
        sku, wh, wht = str(r[0]), str(r[1] or ''), str(r[2] or '')
        avail = int(r[3] or 0)
        if avail <= 0:
            continue
        p = products.get(sku)
        if not p:
            continue
        # 品类线
        food = p['food']
        pot_line = FC['pot_food'] if food else FC['pot_nonfood']
        conf_line = FC['conf_food'] if food else FC['conf_nonfood']
        turn_line = FC['turn_food'] if food else FC['turn_nonfood']
        rate_line = FC['rate_food'] if food else FC['rate_nonfood']
        shelf_m = FC['shelf_food'] if food else FC['shelf_nonfood']
        # 空安全的销售计算
        def safe(fn, d=0.0):
            try: return fn()
            except Exception: return d
        daily = safe(lambda: float(sales_28.get(sku, 0) or 0))
        turnover = round(avail / daily, 1) if daily > 0 else 999.0
        days_zero = 999
        ld = last_order.get(sku, '')
        if ld:
            try: days_zero = (datetime.strptime(ld[:10], "%Y-%m-%d") - today).days * -1
            except Exception: pass
        s90 = sale_90.get(sku)
        s90_days = len(s90['days']) if s90 else 0
        s90_qty = s90['qty'] if s90 else 0
        win_qty = sale_win.get(sku, 0)
        turn_rate = round(win_qty / avail * 100, 1) if avail > 0 else 0
        deep_ratio = round(avail / s90_qty * 100, 1) if s90_qty > 0 else 999.0
        fund = round(avail * p['price'], 0)
        abc = abc_map.get(sku, 'c')
        reason = []
        level = None
        b_storage = None
        # ① 临期风险（有保质期数据）
        bb = p['best_before']
        if bb:
            try:
                dd = (datetime.strptime(bb[:10], "%Y-%m-%d") - today).days
                if dd <= shelf_m * 30:
                    level = 'black'
                    reason.append(f"距保质期{max(dd,0)}天(<{shelf_m}月线临期)")
            except Exception: pass
        # ② B仓仓储费（京东BBCC）
        if wht == 'platform_b' and channel == 'jd':
            days_stored = 0
            if sku in b_arrival:
                days_stored = max((today - b_arrival[sku]).days, 0)
            if days_stored > FC['b_free']:
                vol_m3 = round(avail * p['volume'], 3)
                fee = round(vol_m3 * (days_stored - FC['b_free']) * FC['fee_rate'], 2)
                b_storage = {"days_stored": days_stored, "free_days": FC['b_free'], "volume_m3": vol_m3, "fee_est": fee, "fee_rate": FC['fee_rate']}
                reason.append(f"B仓在库{days_stored}天超{FC['b_free']}天免费期(日仓储费约¥{fee})")
                if level is None and days_stored > FC['b_free'] + 7:
                    level = 'deep'
        # ③ 深度积压
        if level is None and days_zero >= FC['deep_noorder'] and deep_ratio > FC['deep_ratio']:
            level = 'deep'
            reason.append(f"深度积压: 库存占销量{deep_ratio}% 且 {days_zero}天无订单")
        # ④ 确认滞销（零销售超品类确认线 且 存销比超阈值）
        if level is None and days_zero >= conf_line:
            sales_ratio = round(avail / (daily or 1), 1) if daily > 0 else 999.0
            if sales_ratio > FC['ratio_confirm']:
                level = 'confirm'
                reason.append(f"连续{days_zero}天零销售 且 存销比{sales_ratio} 超{FC['ratio_confirm']}")
        # ⑤ 潜在滞销（零销售超潜在线 且 周转超品类线 或 动销率低于线）
        if level is None and days_zero >= pot_line:
            if turnover > turn_line:
                level = 'potential'
                reason.append(f"潜在滞销: {days_zero}天无销售 周转{turnover}天超{turn_line}天线")
            elif turn_rate < rate_line:
                level = 'potential'
                reason.append(f"潜在滞销: 动销率{turn_rate}% 低于{rate_line}%线(近{FC['window']}天)")
        # ⑥ 预警（零销售超预警线）
        if level is None and days_zero >= FC['alert_days']:
            level = 'warn'
            reason.append(f"预警: 连续{days_zero}天零销售")
        # ABC 加权 + 伪滞销
        if level is not None:
            if abc == 'a' and LEVEL_ORDER.get(level, 9) < 3:
                level = 'potential'  # A类核心单品, 不轻易确认/深度处置
                reason.append(f"A类核心单品, 降级为潜在")
            elif abc == 'c' and level == 'warn':
                level = 'potential'  # C类提高一级监控
                reason.append(f"C类小单品, 升级为潜在")
        if level is None:
            continue
        if sku in pseudo and level in ('confirm', 'deep', 'black'):
            reason.append("近14天有补货建议(疑似缺货伪滞销, 建议先核实库存)")
        suggestions.append({
            "sku": sku, "product_name": p['name'], "channel": channel,
            "warehouse": wh, "warehouse_type": wht,
            "stock": avail, "turnover_days": turnover, "fund_occupied": fund,
            "daily_sales": round(daily, 1),
            "days_zero": days_zero, "turnover_rate": turn_rate, "deep_ratio": deep_ratio,
            "abc": abc, "food": food, "best_before": bb,
            "level": level, "reason": reason, "suggestion": SUG.get(level, ''),
            "b_storage": b_storage,
            "disposed": (sku, wh) in disposed, "disposed_action": disposed.get((sku, wh), ''),
        })
    suggestions.sort(key=lambda x: (LEVEL_ORDER.get(x['level'], 9), -x['turnover_days']))
    return ok(suggestions)


@router.get('/export-slow-moving')
def export_slow_moving_excel(channel: str = 'jd', db = get_db()):
    """导出滞销预警为 Excel"""
    from openpyxl import Workbook
    from io import BytesIO
    from fastapi.responses import Response
    from urllib.parse import quote

    result = get_slow_moving_products(db)
    # 按渠道过滤
    import json
    data = result.get("data") if isinstance(result, dict) and "data" in result else (result if isinstance(result, list) else [])
    if channel != 'all':
        products = set()
        try:
            from app.core.database import get_conn as _gconn
            for _r in _gconn().execute("SELECT sku FROM products WHERE channel=?", (channel,)).fetchall():
                products.add(_r[0])
        except Exception:
            products = set()
        data = [x for x in data if x['sku'] in products]
    slow = [x for x in data if x.get('level') != '正常']

    wb = Workbook()
    ws = wb.active
    ws.title = "滞销预警"
    headers = ["SKU","商品","最近下单","天数","库存","级别"]
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    hf = PatternFill(start_color="1d4ed8", end_color="1d4ed8", fill_type="solid")
    hfn = Font(bold=True,color="ffffff",size=11)
    thin = Border(left=Side(style='thin',color='e2e8f0'),right=Side(style='thin',color='e2e8f0'),top=Side(style='thin',color='e2e8f0'),bottom=Side(style='thin',color='e2e8f0'))
    ws.append(headers)
    for c in ws[1]: c.fill=hf; c.font=hfn; c.alignment=Alignment(horizontal='center'); c.border=thin
    for r in slow:
        ws.append([r.get('sku',''),r.get('product_name',''),r.get('last_order_date',''),r.get('days_since_last',0),r.get('stock',0),r.get('level','')])
        for c in ws[ws.max_row]: c.border=thin; c.alignment=Alignment(horizontal='center')
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return Response(content=buf.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":f"attachment; filename*=UTF-8''slow_moving_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"})


@router.get('/summary')
def get_insight_summary(db = get_db()):
    inv = []
    try:
        from app.core.database import get_conn as _gconn
        inv = [{"sku": r[0], "available_qty": r[1], "safety_qty": r[2] or 0}
               for r in _gconn().execute("SELECT sku, available_qty, safety_qty FROM inventory").fetchall()]
    except Exception:
        inv = []
    total = len(inv)
    low_stock = len([x for x in inv if int(x.get("available_qty") or 0) < int(x.get("safety_qty") or 0)])
    out_of_stock = len([x for x in inv if int(x.get("available_qty") or 0) == 0])

    replen_raw = get_replenishment_suggestions(db=db)
    replen = replen_raw.get("data") if isinstance(replen_raw, dict) and "data" in replen_raw else replen_raw
    urgent = len([x for x in replen if x.get("suggested_qty", 0) > 0]) if isinstance(replen, list) else 0

    slow = get_slow_moving_products(db)
    slow_list = slow.get("data") if isinstance(slow, dict) and "data" in slow else (slow if isinstance(slow, list) else [])
    slow_count = len([x for x in slow_list if x.get("level") == "滞销"])
    cold_count = len([x for x in slow_list if x.get("level") == "冷淡"])

    return ok({
        "total_products": total,
        "low_stock": low_stock,
        "out_of_stock": out_of_stock,
        "urgent_replenish": urgent,
        "suggestions_count": len(replen) if isinstance(replen, list) else 0,
        "slow_moving": slow_count,
        "cold_count": cold_count,
    })


@router.get('/trend-analysis')
def trend_analysis(days: int = 30, channel: str = 'jd', db = get_db()):
    """趋势分析：日/周/月维度聚合"""
    from collections import defaultdict
    from datetime import datetime, timedelta
    cutoff = (datetime.utcnow() - timedelta(days=days)).isoformat()
    orders = []
    inventory = []
    try:
        from app.core.database import get_conn as _gconn
        _c = _gconn()
        orders = [{"ordered_at": r[0], "total_amount": r[1], "product_name": r[2]}
                  for r in _c.execute("SELECT ordered_at, total_amount, product_name FROM orders WHERE channel=? AND ordered_at>=?", (channel, cutoff)).fetchall()]
        inventory = [{"available_qty": r[0], "safety_qty": r[1]}
                     for r in _c.execute("SELECT available_qty, safety_qty FROM inventory WHERE channel=?", (channel,)).fetchall()]
    except Exception:
        orders, inventory = [], []

    daily = defaultdict(lambda: {'gmv': 0, 'orders': 0})
    cat_count = defaultdict(int)
    for o in orders:
        date = (o.get('ordered_at') or '')[:10]
        daily[date]['gmv'] += float(o.get('total_amount') or 0)
        daily[date]['orders'] += 1
        cat = o.get('product_name', '未知')[:4]
        cat_count[cat] += 1

    trend = [{'date': d, **v} for d, v in sorted(daily.items())[-days:]]
    cat_pie = [{'name': k, 'value': v} for k, v in sorted(cat_count.items(), key=lambda x: -x[1])[:10]]
    inv_status = {
        'normal': sum(1 for i in inventory if int(i.get('available_qty') or 0) >= int(i.get('safety_qty') or 0)),
        'low': sum(1 for i in inventory if 0 < int(i.get('available_qty') or 0) < int(i.get('safety_qty') or 0)),
        'out': sum(1 for i in inventory if int(i.get('available_qty') or 0) <= 0),
    }
    return {'daily': trend, 'categories': cat_pie, 'inventory_health': inv_status,
            'total_gmv': sum(d['gmv'] for d in trend), 'total_orders': sum(d['orders'] for d in trend)}

@router.get('/anomaly-tracking')
def anomaly_tracking(db = get_db()):
    """异常追踪：告警 + 质量日志汇总"""
    alerts = db.table("alerts").select("*").order("id", desc=True).limit(100).execute().data or []
    quality = db.table("quality_logs").select("*").order("id", desc=True).limit(100).execute().data or []
    events = db.table("events").select("*").order("id", desc=True).limit(100).execute().data or []
    return {
        'alerts': alerts,
        'quality_logs': quality,
        'events': events,
        'summary': {
            'alert_count': len(alerts),
            'active_alerts': sum(1 for a in alerts if a.get('status') == 'active'),
            'error_count': sum(1 for q in quality if q.get('level') == 'error'),
            'event_count': len(events),
        }
    }

@router.post('/sync-from-orders')
def sync_inventory_from_orders(db = get_db(), limit: int = 200):
    """根据最近订单自动调整库存（异步调用）"""
    orders = db.table("orders").select("*").order("id", desc=True).limit(limit).execute().data
    count = 0
    for o in orders:
        try:
            auto_adjust_inventory(o, 'cleansing', db)
            count += 1
        except Exception:
            pass
    return {'ok': True, 'synced': count, 'scanned': len(orders)}


def auto_adjust_inventory(order_data: dict, order_type: str, db):
    sku = order_data.get("sku", "")
    qty = int(float(order_data.get("quantity", 0)))
    if not sku or qty <= 0:
        return

    inv_list = db.table("inventory").select("*").eq("sku", sku).execute().data
    if inv_list:
        inv = inv_list[0]
        avail = int(inv.get("available_qty") or 0)
        if order_type in ("jd_purchase", "cleansing_purchase"):
            new_avail = avail + qty
            db.table("inventory").update({"available_qty": new_avail}).eq("id", inv["id"]).execute()
            inv["available_qty"] = new_avail
        elif order_type in ("sales", "jd_sales", "cleansing"):
            new_avail = max(0, avail - qty)
            db.table("inventory").update({"available_qty": new_avail}).eq("id", inv["id"]).execute()
            inv["available_qty"] = new_avail
        else:
            return
        # Emit inventory.changed so alert/event handlers fire
        try:
            from app.core.events import bus
            bus.emit('inventory.changed', {
                'inventory': inv,
                'action': 'auto_adjust',
                'quantity': qty,
                'order_type': order_type,
            })
        except Exception:
            pass
    else:
        db.table("inventory").insert({
            "sku": sku,
            "product_name": order_data.get("product_name", ""),
            "store": order_data.get("store", ""),
            "available_qty": qty if order_type in ("jd_purchase", "cleansing_purchase") else 0,
            "locked_qty": 0,
            "in_transit_qty": 0,
            "safety_qty": 10,
        }).execute()
@router.get('/with-sales')
def inventory_with_sales(wh_type: str = 'own', channel: str = 'jd', page: int = 0, page_size: int = 0, db = get_db()):
    """库存列表 + 日销 + 在库周转 + 当月出入库
    wh_type: own=自有仓, platform=平台仓(C仓), platform_b=B仓
    page/page_size: 翻页参数，传 0 返回全部
    """
    # 结果缓存 30s（版本号校验，数据变更自动失效）
    import time as _t
    _cache_key = f"{wh_type}|{channel}"
    _now_ts = _t.time()
    try:
        from app.core.dashboard_cache import check_db_version
        _ver = check_db_version()
    except Exception:
        _ver = 0
    _cached = _with_sales_cache.get(_cache_key)
    if _cached and _cached.get('ver') == _ver and _now_ts - _cached.get('ts', 0) < 30:
        return ok(_cached['data'])
    # 惰性归档：每天最多检查一次是否有超期订单需归档（不依赖凌晨任务）
    try:
        _arc = db.table("replenishment_config").select("*").eq("key", "_last_archive_check").execute().data
        _last_arc = _arc[0]['value'] if _arc else ''
        from datetime import timedelta as _td
        if _last_arc != (datetime.utcnow() - _td(days=1)).strftime('%Y-%m-%d'):
            from app.core.scheduler import _task_archive_orders
            _task_archive_orders()
            db.table("replenishment_config").upsert({"key": "_last_archive_check", "value": datetime.utcnow().strftime('%Y-%m-%d'), "channel": "jd", "updated_at": datetime.utcnow().isoformat()}, conflict_col='key')
    except Exception:
        pass
    inv = db.table("inventory").select("*").eq("warehouse_type", wh_type).eq("channel", channel).execute().data or []
    from datetime import datetime, timedelta
    now = datetime.utcnow()
    cur_month = now.strftime('%Y-%m')
    # 日销已从快照读取，orders 全量加载已移除（死代码，改用快照）
    # 出入库记录只取当月
    month_start = now.replace(day=1).strftime('%Y-%m-%d')
    month_end = now.strftime('%Y-%m-%d')
    in_records = db.table("inbound_records").select("*").gte("inbound_date", month_start).eq("channel", channel).execute().data or []
    out_records = db.table("outbound_records").select("*").gte("outbound_date", month_start).eq("channel", channel).execute().data or []
    # 当月出入库汇总
    inbound_month = {}
    for r in in_records:
        s = r['sku']
        inbound_month[s] = inbound_month.get(s, 0) + int(r.get('quantity',0) or 0)
    outbound_month = {}
    for r in out_records:
        s = r['sku']
        outbound_month[s] = outbound_month.get(s, 0) + int(r.get('quantity',0) or 0)
    sales_28 = {}
    products_for_barcode = {}
    try:
        from app.core.database import get_conn as _gconn
        for _r in _gconn().execute("SELECT sku, barcode, price FROM products").fetchall():
            products_for_barcode[_r[0]] = {"sku": _r[0], "barcode": _r[1] or '', "price": _r[2] or 0}
    except Exception:
        products_for_barcode = {}
    # 从快照聚合 28 天日销（替代 orders 全表遍历）
    from app.core.sales_utils import load_daily_sales, calc_sales_multi, rolling_predict
    daily_28 = load_daily_sales(28, db, sku_barcode_map={s: (p.get('barcode','') or '') for s,p in products_for_barcode.items()})
    for key, daily in daily_28.items():
        sales_28[key] = sum(daily.values())
    # 融合日销（一次遍历算 7/14/28 三窗口 + 趋势加权，用于周转天数计算）
    _multi = calc_sales_multi(daily_28, windows=[7, 14, 28])
    _s7, _s14, _s28 = _multi[7], _multi[14], _multi[28]
    _fused = {}
    for _sk in set(list(_s7.keys()) + list(_s14.keys()) + list(_s28.keys())):
        _fused[_sk] = rolling_predict(_s7.get(_sk, 0), _s14.get(_sk, 0), _s28.get(_sk, 0))
    result = []
    # 当查询 B 仓时，预加载 C 仓在途用于 B→C 调拨在途列
    c_transit = {}
    if wh_type == 'platform_b':
        c_inv = db.table("inventory").select("*").eq("warehouse_type", "platform").execute().data or []
        for ci in c_inv:
            s = ci['sku']
            c_transit[s] = c_transit.get(s, 0) + int(ci.get('in_transit_qty', 0) or 0)
    for i in inv:
        sku = i['sku']
        bc = (products_for_barcode.get(sku) or {}).get('barcode', '')
        sales_key = f"{sku}|{bc}" if bc else sku
        ds = round(sales_28.get(sales_key, 0), 1)
        avail = int(i.get('available_qty',0) or 0)
        begin = avail - inbound_month.get(sku, 0) + outbound_month.get(sku, 0)
        # 单价：从 products 主表联表获取
        price = 0
        _p = products_for_barcode.get(sku) or {}
        try: price = float(_p.get('price') or 0)
        except Exception: price = 0
        # 周转天数 = 可用库存 / 融合日销（三窗口 3σ 剔除 + 趋势加权，比简单平均更精准）
        fused_ds = _fused.get(sales_key, 0) or _fused.get(sku, 0)
        turnover_days = round(avail / fused_ds, 1) if fused_ds > 0 else None
        result.append({
            'id': i['id'],
            'sku': sku,
            'barcode': bc,
            'product_name': i.get('product_name',''),
            'price': price,
            'store': i.get('store',''),
            'warehouse': i.get('warehouse',''),
            'warehouse_type': i.get('warehouse_type','platform'),
            'channel': i.get('channel', 'jd'),
            'available_qty': avail,
            'in_transit_qty': int(i.get('in_transit_qty',0) or 0),
            'c_transit': c_transit.get(sku, 0) if wh_type == 'platform_b' else 0,
            'daily_sales': ds,
            'month_inbound': inbound_month.get(sku, int(i.get('month_inbound',0) or 0)),
            'month_outbound': outbound_month.get(sku, int(i.get('month_outbound',0) or 0)),
            'beginning_stock': int(i.get('beginning_stock',0) or 0) or begin,
            'month_start': month_start,
            'month_end': month_end,
            'turnover_days': turnover_days,
        })
    total = len(result)
    # 写缓存（30s TTL + 版本号）
    try:
        _with_sales_cache[_cache_key] = {'data': result, 'ts': _t.time(), 'ver': _ver}
    except Exception:
        pass
    if page > 0 and page_size > 0:
        result = result[(page - 1) * page_size: page * page_size]
        return ok({"items": result, "total": total, "page": page, "page_size": page_size})
    return ok(result)