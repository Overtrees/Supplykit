"""采购建议模块 — 从 insights.py 拆出"""
from fastapi import APIRouter
from app.core.database import get_db
from app.core.response import ok
from datetime import datetime
import json, os

router = APIRouter(prefix="/api/insights", tags=["insights"])


@router.get('/purchase')
def get_purchase_suggestions(days: int = 28, mode: str = 'bbcc', channel: str = 'jd', db = get_db()):
    """采购建议：系统总库存视角，含目标周转控制"""
    from datetime import timedelta
    now = datetime.utcnow()

    raw = {r['key']: r['value'] for r in db.table("replenishment_config").select("*").eq("channel", channel).execute().data}
    purchase_lead_time = int(raw.get('purchase_lead_days', '0'))
    moq_default = int(raw.get('moq', '0'))
    purchase_safety_days = float(raw.get('purchase_safety_days', '0'))

    # 活动系数
    season_key = f'season_config_{mode}'
    sv = db.table('replenishment_config').select('*').eq('key', season_key).execute().data
    season_config = json.loads(sv[0]['value']) if sv and sv[0].get('value') else []
    active_factor = 1.0
    for s in season_config:
        if isinstance(s, dict) and s.get('enabled') and float(s.get('factor', 1.0)) > active_factor:
            active_factor = float(s['factor'])

    # 获取 barcode 映射，用于日销复合 key
    products_map = {p["sku"]: p for p in db.table("products").select("*").execute().data}
    sku_barcode_map = {sku: p.get('barcode', '') or '' for sku, p in products_map.items()}

    # 日销：14+28 双窗口融合
    def purchase_calc(win):
        cutoff = (now - timedelta(days=win)).strftime('%Y-%m-%d')
        daily_raw = {}
        for o in db.table("orders").select("*").execute().data:
            sku = o.get("sku", ""); dt = str(o.get("ordered_at", ""))[:10]; q = int(o.get('quantity',0) or 0)
            if not dt >= cutoff or not sku: continue
            key = sku
            bc = sku_barcode_map.get(sku, '')
            if bc: key = f"{sku}|{bc}"
            if key not in daily_raw: daily_raw[key] = {}
            daily_raw[key][dt] = daily_raw[key].get(dt, 0) + q
        result = {}
        for key, daily in daily_raw.items():
            n = len(daily); total = sum(daily.values()); base_avg = total / win
            if n < 3: result[key] = round(base_avg, 1); continue
            all_d = [(now - timedelta(days=i)).strftime('%Y-%m-%d') for i in range(win)]
            vals = [daily.get(d, 0) for d in all_d]
            mean = sum(vals) / win
            th = max(3 * (sum((v-mean)**2 for v in vals) / win) ** 0.5, mean * 1.5)
            w_sum = w_total = 0
            for idx, v in enumerate(reversed(vals)):
                if abs(v-mean) <= th:
                    w = 1.5 if idx >= win-3 else 1.0
                    w_sum += v * w; w_total += w
            result[key] = round(w_sum/w_total, 1) if w_total > 0 else round(base_avg, 1)
        return result

    def get_purchase_sales(sales_dict, sku):
        """按 sku 查询采购日销，优先用复合 key，降级为 sku"""
        bc = sku_barcode_map.get(sku, '')
        if bc:
            val = sales_dict.get(f"{sku}|{bc}")
            if val is not None: return val
        return sales_dict.get(sku, 0)

    sales_14 = purchase_calc(14)
    sales_28 = purchase_calc(28)
    fused_sales = {}
    for sku in set(sku_barcode_map.keys()):
        s14 = get_purchase_sales(sales_14, sku); s28 = get_purchase_sales(sales_28, sku)
        if s14 > s28 * 1.15: w14, w28 = 0.55, 0.45
        elif s14 < s28 * 0.85: w14, w28 = 0.35, 0.65
        else: w14, w28 = 0.20, 0.80
        fused_sales[sku] = round(s14 * w14 + s28 * w28, 1)

    # 系统总库存
    inv_data = db.table("inventory").select("*").eq("channel", channel).execute().data
    stock_by_sku = {}; b_avail = {}
    for i in inv_data:
        s = i['sku']
        if s not in stock_by_sku:
            stock_by_sku[s] = {'available':0,'transit':0,'safety':0,'safety_days':0,
                               'own_avail':0,'own_transit':0,'plat_avail':0,'plat_transit':0,'own_warehouse':''}
            b_avail[s] = 0
        qty = int(i.get('available_qty',0) or 0); tty = int(i.get('in_transit_qty',0) or 0)
        wt = i.get('warehouse_type','platform')
        stock_by_sku[s]['available'] += qty; stock_by_sku[s]['transit'] += tty
        stock_by_sku[s]['safety'] += int(i.get('safety_qty',0) or 0)
        sd = float(i.get('safety_days',0) or 0)
        if sd > stock_by_sku[s]['safety_days']: stock_by_sku[s]['safety_days'] = sd
        if wt == 'platform_b': b_avail[s] += qty
        elif wt == 'own':
            stock_by_sku[s]['own_avail'] += qty; stock_by_sku[s]['own_transit'] += tty
            if not stock_by_sku[s]['own_warehouse']: stock_by_sku[s]['own_warehouse'] = i.get('warehouse','')
        else: stock_by_sku[s]['plat_avail'] += qty; stock_by_sku[s]['plat_transit'] += tty

    products = {p["sku"]: p for p in db.table("products").select("*").execute().data}

    result = []
    for sku, st in stock_by_sku.items():
        ds = round(fused_sales.get(sku, 0) * active_factor, 1)
        sys_total = st['available'] + st['transit']
        safety_days = st['safety_days'] if st['safety_days'] > 0 else purchase_safety_days
        eff_safety = round(ds * safety_days) if ds > 0 else 0
        purchase_qty = max(round(ds * purchase_lead_time) + eff_safety - sys_total, 0) if ds > 0 else 0
        purchase_qty = max(purchase_qty, moq_default) if purchase_qty > 0 else 0
        prod = products.get(sku, {})
        box_qty = int(prod.get('box_qty', 1) or 1)
        actual_purchase = (purchase_qty + box_qty - 1) // box_qty * box_qty if purchase_qty > 0 else 0
        days_to_empty = round(st['available'] / ds, 1) if ds > 0 else 999
        after_stock = st['own_avail'] + st['own_transit'] + actual_purchase
        after_turnover = round(after_stock / ds, 1) if ds > 0 else 999
        target_turn = int(raw.get('max_turnover_days', '0'))
        c_consume = round(ds * purchase_lead_time) if ds > 0 else 0
        note = ""
        if purchase_qty > 0:
            note = f"消耗{c_consume}+安全{eff_safety} -系统总库存{int(sys_total)} ={int(purchase_qty)}"
            note += f" | 箱规{box_qty}件, 实购{actual_purchase}件"
            note += f"（{actual_purchase//box_qty}箱）" if box_qty > 1 else ""
            note += f", 补后周转{after_turnover}天"
            if target_turn > 0:
                note += f" > 目标{target_turn}天" if after_turnover > target_turn else f" < 目标{target_turn}天"

        result.append({
            'sku': sku, 'product_name': prod.get('product_name', ''),
            'store': prod.get('store', ''), 'warehouse': st['own_warehouse'], 'category': prod.get('category', ''),
            'sys_available': st['available'], 'sys_transit': st['transit'], 'sys_total': sys_total,
            'own_available': st['own_avail'], 'own_transit': st['own_transit'],
            'plat_available': st['plat_avail'], 'plat_transit': st['plat_transit'],
            'b_available': b_avail.get(sku, 0),
            'safety_qty': st['safety'], 'daily_sales': ds,
            'daily_sales_14': get_purchase_sales(sales_14, sku), 'daily_sales_28': get_purchase_sales(sales_28, sku),
            'purchase_qty': purchase_qty, 'box_qty': box_qty, 'actual_purchase': actual_purchase,
            'after_stock': st['own_avail'] + purchase_qty, 'after_turnover': after_turnover,
            'target_turnover': target_turn,
            'days_to_empty': days_to_empty, 'note': note,
        })

    result.sort(key=lambda x: x['days_to_empty'])
    # 创建/关闭采购告警
    for r in result:
        if r['purchase_qty'] > 0 and r['days_to_empty'] < 14:
            try:
                ex = db.table("alerts").select("id").eq("alert_type","purchase_need").eq("related_sku",r['sku']).eq("status","active").execute().data
                if not ex:
                    db.table("alerts").insert({"alert_type":"purchase_need","title":f"需采购: {r['product_name']}",
                        "description":f"可用{r['available_qty']}件, 建议采购{r['purchase_qty']}件, 可撑{r['days_to_empty']}天",
                        "severity":"warning","source":"purchase_engine","related_sku":r['sku'],"status":"active"}).execute()
            except: pass
        elif r['purchase_qty'] == 0:
            try:
                db.table("alerts").update({"status":"closed"}).eq("alert_type","purchase_need").eq("related_sku",r['sku']).eq("status","active").execute()
            except: pass
    return ok(result)


@router.get('/export-purchase-suggestions')
def export_purchase_suggestions_excel(days: int = 28, mode: str = 'bbcc', channel: str = 'jd', db = get_db()):
    """导出采购建议为 Excel"""
    from openpyxl import Workbook
    from io import BytesIO
    from fastapi.responses import Response
    from urllib.parse import quote

    data = get_purchase_suggestions(days=days, mode=mode, channel=channel, db=db)
    suggestions = data.get("data") if isinstance(data, dict) and "data" in data else data

    wb = Workbook()
    ws = wb.active
    ws.title = "采购建议"
    headers = ["序号","SKU","商品名称","仓库","系统总库存","系统可用","系统在途",
               "自有可用","自有在途","平台可用","平台在途","B仓可用",
               "日销(融合)","日销14","日销28","建议采购量","箱规","实购数量","补后周转","目标周转","可撑天数","备注"]
    ws.append(headers)

    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    head_fill = PatternFill(start_color="1d4ed8", end_color="1d4ed8", fill_type="solid")
    head_font = Font(bold=True, color="ffffff", size=11)
    thin = Border(left=Side(style='thin',color='e2e8f0'), right=Side(style='thin',color='e2e8f0'),
                  top=Side(style='thin',color='e2e8f0'), bottom=Side(style='thin',color='e2e8f0'))
    for cell in ws[1]:
        cell.fill = head_fill; cell.font = head_font
        cell.alignment = Alignment(horizontal='center'); cell.border = thin

    for i, r in enumerate(suggestions, 1):
        ws.append([i, r["sku"], r["product_name"], r["warehouse"], r["sys_total"], r["sys_available"], r["sys_transit"],
            r["own_available"], r["own_transit"], r["plat_available"], r["plat_transit"], r["b_available"],
            r["daily_sales"], r["daily_sales_14"], r["daily_sales_28"],
            r["purchase_qty"], r["box_qty"], r["actual_purchase"], r["after_turnover"], r["target_turnover"],
            r["days_to_empty"] if r["days_to_empty"] < 999 else "∞", r["note"]])
        for cell in ws[ws.max_row]: cell.border = thin; cell.alignment = Alignment(horizontal='center')

    widths = [6,14,20,12,12,10,10,10,10,10,10,10,10,10,10,12,8,10,10,10,10,30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1,i).column_letter].width = w

    ws2 = wb.create_sheet("汇总")
    ws2.append(["采购建议汇总"]); ws2.append(["生成时间", datetime.utcnow().strftime("%Y-%m-%d %H:%M")])
    ws2.append(["建议采购SKU数", len(suggestions)])
    ws2.append(["建议采购总量", sum(r["purchase_qty"] for r in suggestions)])
    ws2.merge_cells('A1:D1'); ws2['A1'].font = Font(bold=True, size=14)

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"采购建议_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return Response(content=buf.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"})