from fastapi import APIRouter, Depends, UploadFile, File
from app.core.database import get_db
import csv, io
from openpyxl import load_workbook

router = APIRouter(prefix="/api/inventory", tags=["inventory"])

@router.get("")
def list_inventory(db = get_db(), store: str = ""):
    q = db.table("inventory").select("*")
    if store:
        q = q.eq("store", store)
    data = q.order("id", desc=True).execute().data
    return data

@router.post("")
def create_inventory(body: dict, db = get_db()):
    data = db.table("inventory").insert({
        "sku": body.get("sku"),
        "product_name": body.get("product_name"),
        "store": body.get("store", ""),
        "warehouse": body.get("warehouse", ""),
        "available_qty": int(body.get("available_qty", 0)),
        "locked_qty": int(body.get("locked_qty", 0)),
        "in_transit_qty": int(body.get("in_transit_qty", 0)),
        "safety_qty": int(body.get("safety_qty", 10)),
        "status": body.get("status", "active"),
    }).execute().data
    inv = data[0] if data else None
    if inv:
        try:
            from app.core.events import bus
            bus.emit('inventory.changed', {
                'inventory': inv,
                'action': 'create',
                'quantity': inv.get('available_qty'),
            })
        except Exception:
            pass
    return inv or {"ok": True}

@router.put("/{iid}")
def update_inventory(iid: int, body: dict, db = get_db()):
    db.table("inventory").update(body).eq("id", iid).execute()
    inv = db.table("inventory").select("*").eq("id", iid).execute().data
    inv = inv[0] if inv else None
    if inv:
        try:
            from app.core.events import bus
            bus.emit('inventory.changed', {
                'inventory': inv,
                'action': 'update',
                'quantity': inv.get('available_qty'),
            })
        except Exception:
            pass
        # 直接写入事件（不依赖事件总线）
        try:
            from app.api.routes.events import create_event
            create_event(db, 'stock.changed', 'inventory', str(inv['id']),
                         f"库存变动: {inv.get('product_name', inv.get('sku',''))}",
                         {'available_qty': inv.get('available_qty'), 'action': 'update'})
        except Exception:
            pass
        # 直接检查并触发规则
        try:
            from app.core.rules import evaluate
            evaluate('inventory.changed', {'inv': inv, 'db': db, 'sku': inv.get('sku','')})
        except Exception:
            pass
    return {"ok": True}

@router.delete("/{iid}")
def delete_inventory(iid: int, db = get_db()):
    db.table("inventory").delete().eq("id", iid).execute()
    return {"ok": True}

@router.post("/adjust")
def adjust_inventory(body: dict, db = get_db()):
    iid = body.get("id")
    action = body.get("action")
    qty = int(body.get("quantity", 0))
    inv = db.table("inventory").select("*").eq("id", iid).execute().data
    inv = inv[0] if inv else None
    if not inv:
        return {"ok": False, "error": "not found"}
    avail = int(inv.get("available_qty") or 0)
    new_avail = avail
    if action == "in":
        new_avail = avail + qty
        db.table("inventory").update({"available_qty": new_avail}).eq("id", iid).execute()
    elif action == "out":
        new_avail = max(0, avail - qty)
        db.table("inventory").update({"available_qty": new_avail}).eq("id", iid).execute()
    elif action == "set":
        new_avail = qty
        db.table("inventory").update({"available_qty": new_avail}).eq("id", iid).execute()
    
    inv["available_qty"] = new_avail
    return {"ok": True}


@router.post("/import")
def import_inventory(file: UploadFile = File(...), db = get_db()):
    try:
        content = file.file.read()
    except Exception as e:
        return {'ok': False, 'error': f'读取文件失败: {e}', 'imported': 0}
    rows = []
    try:
        if (file.filename or '').endswith('.csv'):
            text = content.decode('utf-8-sig')
            rows = list(csv.DictReader(io.StringIO(text)))
        else:
            wb = load_workbook(io.BytesIO(content))
            ws = wb.active
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: row[i] for i in range(len(headers)) if row[i] is not None})
    except Exception as e:
        return {'ok': False, 'error': f'解析文件失败: {e}', 'imported': 0}
    if not rows:
        return {'ok': False, 'error': '文件内容为空', 'imported': 0}

    TABLE_COLS = {'sku','product_name','store','warehouse',
                  'available_qty','locked_qty','in_transit_qty',
                  'safety_qty','safety_days','raw_data','source','owner_id'}
    ALIAS = {
        'SKU':'sku','商品编号':'sku','商品名称':'product_name','名称':'product_name',
        '店铺':'store','仓库':'warehouse',
        '可用库存':'available_qty','可用':'available_qty',
        '锁定库存':'locked_qty','锁定':'locked_qty',
        '在途':'in_transit_qty','在途库存':'in_transit_qty',
        '安全线':'safety_qty','安全库存':'safety_qty','安全天数':'safety_days',
    }

    inserted = 0
    duplicates = 0
    skipped = 0
    imported_cols = set()
    for row in rows:
        mapped = {}
        raw_extra = {}
        file_provided = set()
        for raw_col, raw_val in row.items():
            if raw_col is None: continue
            alias = ALIAS.get(raw_col.strip())
            if alias and alias in TABLE_COLS:
                mapped[alias] = str(raw_val).strip() if raw_val is not None else ''
                file_provided.add(alias)
            elif raw_col.strip() in TABLE_COLS:
                mapped[raw_col.strip()] = str(raw_val).strip() if raw_val is not None else ''
                file_provided.add(raw_col.strip())
            else:
                raw_extra[raw_col.strip()] = str(raw_val).strip() if raw_val is not None else ''
        if not mapped.get('sku'):
            skipped += 1
            continue
        mapped['available_qty'] = int(float(mapped.get('available_qty') or 0))
        mapped['locked_qty'] = int(float(mapped.get('locked_qty') or 0))
        mapped['in_transit_qty'] = int(float(mapped.get('in_transit_qty') or 0))
        mapped['safety_qty'] = int(float(mapped.get('safety_qty') or 10))
        mapped['safety_days'] = float(mapped.get('safety_days') or 0)

        # 查重
        sk = mapped.get('sku','')
        existing = db.table("inventory").select("*").eq("sku", sk).eq("store", mapped.get('store','')).eq("warehouse", mapped.get('warehouse','')).execute().data
        is_dup = len(existing) > 0
        if is_dup and existing:
            existing_row = existing[0]
            for col in list(mapped.keys()):
                if col in ('sku', 'source'): continue
                if col not in file_provided:
                    mapped[col] = existing_row.get(col, mapped.get(col))
                else:
                    v = mapped[col]
                    ev = existing_row.get(col)
                    if isinstance(v, str) and not v.strip():
                        mapped[col] = ev if ev is not None else mapped[col]
                    elif isinstance(v, (int, float)) and v == 0 and ev:
                        mapped[col] = ev
            duplicates += 1
        else:
            inserted += 1

        mapped['source'] = 'import'
        if raw_extra:
            mapped['raw_data'] = json.dumps(raw_extra, ensure_ascii=False)
        imported_cols.update(k for k in mapped.keys() if k not in ('raw_data','source'))

        db.table("inventory").upsert(mapped)
    from app.core.events import bus
    bus.emit('inventory.imported', {'count': inserted})
    return {
        'ok': True, 'imported': inserted, 'duplicates': duplicates, 'from_file': file.filename,
        'total_rows': len(rows), 'skipped': skipped,
        'columns_mapped': list(imported_cols),
    }
