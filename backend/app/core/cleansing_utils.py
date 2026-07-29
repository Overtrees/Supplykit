"""清洗工具模块 — 文件解析 + 字段清洗逻辑"""
import csv, io, re, json
from openpyxl import load_workbook


def parse_file(content: bytes, filename: str):
    """解析上传文件，返回二维数组（第一行为表头）"""
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    if ext == 'csv':
        try:
            raw = content.decode('utf-8-sig')
        except:
            raw = content.decode('gbk', errors='replace')
        reader = csv.DictReader(io.StringIO(raw))
        headers = reader.fieldnames or []
        rows = [list(row.values()) for row in reader]
        return [headers] + rows if headers else []
    elif ext in ('xlsx', 'xls'):
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            processed = [str(c) if c is not None else '' for c in row]
            rows.append(processed)
        return rows
    else:
        return []


def cleanse_value(raw_val, cfg):
    """根据配置清洗单个字段值"""
    val = str(raw_val).strip() if raw_val is not None else ''
    vtype = cfg.get('type', 'string')
    default = cfg.get('default', '')
    if not val or val == '':
        return default
    if vtype == 'number':
        cleaned = re.sub(r'[^\d.\-]', '', val)
        try:
            return float(cleaned) if '.' in cleaned else int(float(cleaned))
        except:
            return 0
    elif vtype == 'date':
        fmt = cfg.get('format', '')
        if fmt == 'YMD':
            return val[:10]
        return val[:10]
    elif vtype == 'int':
        cleaned = re.sub(r'[^\d\-]', '', val)
        try:
            return int(float(cleaned))
        except:
            return 0
    return val


def extract_field_mapping(headers):
    """自动检测字段映射"""
    rules = {
        '订单号': 'order_no', 'order_no': 'order_no', 'orderno': 'order_no',
        '商品编号': 'sku', 'sku': 'sku', '货号': 'sku', '编码': 'sku',
        '商品名称': 'product_name', 'product_name': 'product_name', '品名': 'product_name', '名称': 'product_name',
        '店铺': 'store', 'store': 'store', '门店': 'store',
        '仓库': 'warehouse', 'warehouse': 'warehouse', '库房': 'warehouse',
        '数量': 'quantity', 'quantity': 'quantity', 'qty': 'quantity',
        '单价': 'unit_price', 'unit_price': 'unit_price', '价格': 'unit_price', 'price': 'unit_price',
        '金额': 'total_amount', 'total_amount': 'total_amount', '总额': 'total_amount',
        '状态': 'order_status', 'order_status': 'order_status', '订单状态': 'order_status',
        '时间': 'ordered_at', 'ordered_at': 'ordered_at', '下单时间': 'ordered_at', '日期': 'ordered_at',
        '平台': 'platform', 'platform': 'platform',
        '可用': 'available_qty', 'available_qty': 'available_qty', '库存': 'available_qty',
        '在途': 'in_transit_qty', 'in_transit_qty': 'in_transit_qty',
        '安全库存': 'safety_qty', 'safety_qty': 'safety_qty', '安全': 'safety_qty',
    }
    mapping = {}
    for h in headers:
        key = h.strip().lower().replace('（', '(').replace('）', ')')
        if key in rules:
            mapping[h] = rules[key]
    return mapping


def detect_target(headers, mapping):
    """自动检测导入目标表类型"""
    fields = set(mapping.values())
    if 'order_no' in fields or 'ordered_at' in fields:
        return 'orders'
    if 'in_transit_qty' in fields:
        return 'inventory'
    return 'orders'