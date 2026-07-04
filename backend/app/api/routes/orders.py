from fastapi import APIRouter, Depends, UploadFile, File
from datetime import datetime
import json
import csv
import io
from openpyxl import load_workbook
from app.core.database import get_db

router = APIRouter(prefix="/api/orders", tags=["orders"])

@router.get("")
def list_orders(db = get_db(), page: int = 1, page_size: int = 50, search: str = '', status: str = '', store: str = '', sort_by: str = 'id', sort_order: str = 'desc'):
    all_rows = db.table("orders").select("*").execute().data
    filtered = []
    for row in all_rows:
        if search:
            s = search.lower()
            if s not in (row.get('order_no','') or '').lower() and s not in (row.get('product_name','') or '').lower() and s not in (row.get('sku','') or '').lower():
                continue
        if status and row.get('order_status') != status: continue
        if store and row.get('store') != store: continue
        filtered.append(row)
    total = len(filtered)
    desc = sort_order == 'desc'
    filtered.sort(key=lambda r: (r.get(sort_by) or 0) if isinstance(r.get(sort_by), (int,float)) else str(r.get(sort_by,'')), reverse=desc)
    start = (page - 1) * page_size
    items = filtered[start:start + page_size]
    return {'total': total, 'page': page, 'page_size': page_size, 'total_pages': max(1, (total + page_size - 1) // page_size), 'items': items}

@router.post('/batch-delete')
def batch_delete_orders(ids: str = '', db = get_db()):
    if not ids or ids == 'auto':
        data = db.table("orders").delete().ilike("order_no", "AUTO-%").execute().data
        deleted = len(data)
    else:
        id_list = [int(x.strip()) for x in ids.split(',') if x.strip().isdigit()]
        data = db.table("orders").delete().in_("id", id_list).execute().data
        deleted = len(data)
    return {'ok': True, 'deleted': deleted}

@router.delete('/{oid}')
def delete_order(oid: int, db = get_db()):
    db.table("orders").delete().eq("id", oid).execute()
    return {'ok': True}

@router.post('/import')
def import_orders(file: UploadFile = File(...), db = get_db()):
    try:
        content = file.file.read()
    except Exception as e:
        return {'ok': False, 'error': f'读取文件失败: {e}', 'imported': 0}

    rows = []
    try:
        if (file.filename or '').endswith('.csv'):
            text = content.decode('utf-8-sig')
            rows = list(csv.DictReader(io.StringIO(text)))
        else:
            wb = load_workbook(io.BytesIO(content))
            ws = wb.active
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: row[i] for i in range(len(headers)) if row[i] is not None})
    except Exception as e:
        return {'ok': False, 'error': f'解析文件失败: {e}', 'imported': 0}
    if not rows:
        return {'ok': False, 'error': '文件内容为空', 'imported': 0}

    TABLE_COLS = {'order_no','store','warehouse','sku','product_name',
                  'quantity','unit_price','total_amount','data_source',
                  'order_status','ordered_at','platform','supplier','remark',
                  'parent_order_no','raw_data','source','owner_id'}
    ALIAS = {
        '订单号':'order_no','订单编号':'order_no','采购单号':'order_no',
        '商品编号':'sku','货号':'sku','SKU':'sku','sku':'sku',
        '商品名称':'product_name','产品名称':'product_name','名称':'product_name',
        '数量':'quantity','采购数量':'quantity','订货数量':'quantity',
        '单价':'unit_price','价格':'unit_price','采购价格':'unit_price',
        '金额':'total_amount','总金额':'total_amount','采购金额':'total_amount','实收金额':'total_amount',
        '店铺':'store','店铺名':'store','门店':'store',
        '仓库':'warehouse','京东仓库':'warehouse','发货仓':'warehouse','配送中心':'warehouse',
        '状态':'order_status','订单状态':'order_status',
        '日期':'ordered_at','订购时间':'ordered_at','下单时间':'ordered_at','入库时间':'ordered_at',
        '平台':'platform','订单来源':'platform','来源':'platform','采购渠道':'platform',
        '供应商':'supplier','供应商名称':'supplier',
        '备注':'remark','订货备注':'remark',
    }

    inserted = 0
    imported_items = []
    skipped = 0
    duplicates = 0
    for row in rows:
        mapped = {}
        raw_extra = {}
        file_provided = set()  # 记录文件实际提供了哪些 DB 字段
        for raw_col, raw_val in row.items():
            if raw_col is None: continue
            alias = ALIAS.get(raw_col.strip())
            if alias and alias in TABLE_COLS:
                mapped[alias] = str(raw_val).strip() if raw_val is not None else ''
                file_provided.add(alias)
            elif raw_col.strip() in TABLE_COLS:
                mapped[raw_col.strip()] = str(raw_val).strip() if raw_val is not None else ''
                file_provided.add(raw_col.strip())
            else:
                raw_extra[raw_col.strip()] = str(raw_val).strip() if raw_val is not None else ''

        if not mapped.get('order_no'):
            skipped += 1
            continue

        try:
            mapped['quantity'] = int(float(mapped.get('quantity') or 0))
            mapped['unit_price'] = float(mapped.get('unit_price') or 0)
            mapped['total_amount'] = float(mapped.get('total_amount') or 0)
        except ValueError:
            skipped += 1
            continue

        # 查重：同 order_no + sku 已存在则为重复
        o, s = mapped.get('order_no',''), mapped.get('sku','')
        existing = db.table("orders").select("*").eq("order_no", o).eq("sku", s).execute().data
        is_dup = len(existing) > 0

        if is_dup and existing:
            # 重复记录→智能合并：只覆盖文件实际提供的字段，其余保持数据库原值
            existing_row = existing[0]
            for col in list(mapped.keys()):
                if col in ('order_no', 'sku', 'data_source'):
                    continue  # 键字段和来源标记始终使用文件值
                if col not in file_provided:
                    # 文件没提供该列 → 保持原值
                    mapped[col] = existing_row.get(col, mapped.get(col))
                else:
                    # 文件提供了该列但值为空/0 → 也保持原值
                    imported_val = mapped[col]
                    existing_val = existing_row.get(col)
                    if isinstance(imported_val, str) and not imported_val.strip():
                        mapped[col] = existing_val if existing_val is not None else mapped[col]
                    elif isinstance(imported_val, (int, float)) and imported_val == 0 and existing_val:
                        mapped[col] = existing_val

        mapped['data_source'] = 'import'
        if raw_extra:
            mapped['raw_data'] = json.dumps(raw_extra, ensure_ascii=False)

        db.table("orders").upsert(mapped)
        if is_dup:
            duplicates += 1
        else:
            inserted += 1
        imported_items.append(mapped)

    from app.core.events import bus
    bus.emit('order.imported', {'count': inserted})
    if imported_items:
        bus.emit('order.created', {'items': imported_items, 'order_type': 'import'})
    return {
        'ok': True, 'imported': inserted, 'duplicates': duplicates, 'from_file': file.filename,
        'total_rows': len(rows), 'skipped': skipped,
        'columns_mapped': [k for k in (imported_items[0] if imported_items else {}).keys() if k != 'raw_data'],
        'columns_rawdata': list(raw_extra.keys()) if imported_items and raw_extra else [],
    }
